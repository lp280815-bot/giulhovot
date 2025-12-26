import io
import os
import logging
import uuid
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from pathlib import Path
from collections import defaultdict
from datetime import datetime, timezone, timedelta
from typing import List, Optional
from urllib.parse import quote

import openpyxl
from openpyxl.styles import PatternFill, Alignment, Font
import requests
import aiofiles

from fastapi import FastAPI, APIRouter, UploadFile, File, HTTPException, Form
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from starlette.responses import RedirectResponse
from motor.motor_asyncio import AsyncIOMotorClient
from pydantic import BaseModel, Field, ConfigDict
import httpx


ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Microsoft OAuth2 Configuration
MICROSOFT_CLIENT_ID = os.environ.get('MICROSOFT_CLIENT_ID', '')
MICROSOFT_TENANT_ID = os.environ.get('MICROSOFT_TENANT_ID', '')
MICROSOFT_CLIENT_SECRET = os.environ.get('MICROSOFT_CLIENT_SECRET', '')
MICROSOFT_REDIRECT_URI = os.environ.get('MICROSOFT_REDIRECT_URI', '')
FRONTEND_URL = os.environ.get('FRONTEND_URL', 'http://localhost:3000')
MICROSOFT_AUTHORITY = f"https://login.microsoftonline.com/{MICROSOFT_TENANT_ID}"
MICROSOFT_SCOPES = ["https://graph.microsoft.com/Mail.Send", "https://graph.microsoft.com/User.Read", "offline_access"]

# Create the main app
app = FastAPI()
api_router = APIRouter(prefix="/api")

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# ========= N8N Configuration =========
N8N_WEBHOOK_URL = "https://riseelena.app.n8n.cloud/webhook/giulhovot-ijeaen7cc8ruf8ncxnr88d"

# ========= Colors Configuration =========
GREEN_RGB = "FF00FF00"
ORANGE_RGB = "FFFFA500"
PURPLE_RGB = "FFCC99FF"
BLUE_RGB = "FFADD8E6"

GREEN_FILL = PatternFill(start_color=GREEN_RGB, end_color=GREEN_RGB, fill_type="solid")
ORANGE_FILL = PatternFill(start_color=ORANGE_RGB, end_color=ORANGE_RGB, fill_type="solid")
PURPLE_FILL = PatternFill(start_color=PURPLE_RGB, end_color=PURPLE_RGB, fill_type="solid")
BLUE_FILL = PatternFill(start_color=BLUE_RGB, end_color=BLUE_RGB, fill_type="solid")


# ========= Pydantic Models =========
class StatusCheck(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    client_name: str
    timestamp: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class StatusCheckCreate(BaseModel):
    client_name: str

class Supplier(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    account_number: str
    name: str
    currency: Optional[str] = "ש\"ח"
    vat_number: Optional[str] = None
    purchase_account: Optional[str] = None
    purchase_account_desc: Optional[str] = None
    email: Optional[str] = None
    phone: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class SupplierCreate(BaseModel):
    account_number: str
    name: str
    currency: Optional[str] = "ש\"ח"
    vat_number: Optional[str] = None
    purchase_account: Optional[str] = None
    purchase_account_desc: Optional[str] = None
    email: Optional[str] = None
    phone: Optional[str] = None

class SupplierUpdate(BaseModel):
    account_number: Optional[str] = None
    name: Optional[str] = None
    currency: Optional[str] = None
    vat_number: Optional[str] = None
    purchase_account: Optional[str] = None
    purchase_account_desc: Optional[str] = None
    email: Optional[str] = None
    phone: Optional[str] = None

class SupplierImportResult(BaseModel):
    total: int = 0
    imported: int = 0
    updated: int = 0
    errors: int = 0
    error_messages: List[str] = []

class ProcessingHistory(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    filename: str
    processed_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    green_matches: int = 0
    orange_matches: int = 0
    purple_matches: int = 0
    blue_matches: int = 0
    total_rows: int = 0
    status: str = "completed"

class N8NTriggerRequest(BaseModel):
    client_name: str

class ProcessingStats(BaseModel):
    green_matches: int = 0
    orange_matches: int = 0
    purple_matches: int = 0
    blue_matches: int = 0
    special_treatment: int = 0  # Rows that didn't match any category
    command_rows: int = 0  # Rows marked for "לעשות פקודה"
    total_rows: int = 0
    emails_generated: int = 0

class MoveRowRequest(BaseModel):
    row_index: int
    from_category: str
    to_category: str
    row_data: dict

class DetailedRow(BaseModel):
    account: str = ""
    account_description: str = ""  # תאור חשבון
    supplier_account: str = ""  # חש. ספק
    name: str = ""
    amount: float = 0
    date: str = ""
    details: str = ""
    invoice: str = ""

class ProcessingDetails(BaseModel):
    green: List[DetailedRow] = []
    orange: List[DetailedRow] = []
    purple: List[DetailedRow] = []
    blue: List[DetailedRow] = []
    emails: List[DetailedRow] = []
    special: List[DetailedRow] = []

class SendEmailRequest(BaseModel):
    sender_email: str
    sender_password: str
    recipient_email: str
    subject: str
    body: str
    sender_name: Optional[str] = None


# ========= Helper Functions =========
def parse_amount(val):
    """Convert value to numeric amount (float) with handling for empty and thousands separators."""
    if val is None or val == "":
        raise ValueError("empty")
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    if s == "":
        raise ValueError("empty")
    s = s.replace(",", "")
    return float(s)


def detect_headers(ws):
    """Detect header row: tries row 1 then row 2."""
    candidates = [1, 2]
    chosen_row = None
    headers = {}

    for row_idx in candidates:
        row_cells = ws[row_idx]
        row_values = [str(c.value).strip() if c.value is not None else "" for c in row_cells]
        if any(v for v in row_values):
            tmp_headers = {str(c.value).strip(): c.column for c in ws[row_idx] if c.value}
            if "חשבון" in tmp_headers and "חוב לחשבונית" in tmp_headers:
                chosen_row = row_idx
                headers = tmp_headers
                break
            if not headers:
                chosen_row = row_idx
                headers = tmp_headers

    if chosen_row is None:
        chosen_row = 1
        headers = {str(c.value).strip(): c.column for c in ws[1] if c.value}

    return chosen_row, headers


def cell_rgb(cell):
    try:
        return cell.fill.start_color.rgb
    except Exception:
        return None


def has_any_color(cell):
    """Check if cell has one of the logic colors."""
    return cell.fill.fill_type == "solid" and cell_rgb(cell) in {
        GREEN_RGB, ORANGE_RGB, PURPLE_RGB, BLUE_RGB
    }


def ensure_summary_sheet(wb, title, counts):
    """Create/clean summary sheet and populate data."""
    if title in wb.sheetnames:
        ws_sum = wb[title]
        for row in ws_sum.iter_rows():
            for c in row:
                c.value = None
    else:
        ws_sum = wb.create_sheet(title)

    ws_sum["A1"] = "מס ספק"
    ws_sum["B1"] = "כמות שורות מותאמות"

    r = 2
    for acc, cnt in counts.items():
        if acc is None or cnt <= 0:
            continue
        ws_sum.cell(r, 1, acc)
        ws_sum.cell(r, 2, cnt)
        r += 1


def build_email_mapping(helper_wb):
    """Build dictionary {account/supplier name -> email} from helper file."""
    ws_help = helper_wb.active
    header_row, headers = detect_headers(ws_help)

    col_acc = headers.get("חשבון") or headers.get("מס ספק")
    col_name = headers.get("שם ספק") or headers.get("תאור חשבון") or headers.get("תיאור חשבון")
    col_email = (
        headers.get("מייל") or headers.get("מייל ספק") or 
        headers.get("Email") or headers.get("E-mail")
    )

    email_map = {}

    if col_email is None:
        return email_map

    if col_acc is not None:
        for row in ws_help.iter_rows(min_row=header_row + 1):
            acc = row[col_acc - 1].value
            email = row[col_email - 1].value
            if acc and email:
                email_map[str(acc).strip()] = str(email).strip()

    if col_name is not None:
        for row in ws_help.iter_rows(min_row=header_row + 1):
            name = row[col_name - 1].value
            email = row[col_email - 1].value
            if name and email:
                email_map[str(name).strip()] = str(email).strip()

    return email_map


def process_workbook(wb, email_mapping=None):
    """Run all logics 1-7 on the workbook."""
    ws = wb.active
    header_row, headers = detect_headers(ws)

    col_acc = headers.get("חשבון")
    col_amt = headers.get("חוב לחשבונית")
    col_type = headers.get("סוג תנועה")
    col_name = headers.get("תאור חשבון") or headers.get("שם ספק") or headers.get("תיאור חשבון")
    col_pay = headers.get("תאריך תשלום")

    if col_acc is None or col_amt is None:
        raise ValueError("לא נמצאו עמודות 'חשבון' ו/או 'חוב לחשבונית'.")

    if col_name is None:
        col_name = 3
    if col_pay is None:
        col_pay = 4

    # Try to find document number column
    col_doc = headers.get("מס' אסמכתא") or headers.get("מספר אסמכתא") or headers.get("אסמכתא") or headers.get("מס' מסמך")

    data_start_row = header_row + 1
    company_name = ws["C1"].value if ws["C1"].value is not None else ""

    # Try to find additional columns
    col_details = headers.get("פרטים") or headers.get("פירוט") or headers.get("תאור")
    col_invoice = headers.get("חשבונית") or headers.get("מס' חשבונית") or headers.get("מספר חשבונית") or headers.get("אסמכתא")
    col_acc_desc = headers.get("תאור חשבון") or headers.get("תיאור חשבון") or headers.get("שם חשבון")
    col_supplier_acc = headers.get("חש. ספק") or headers.get("חשבון ספק") or headers.get("מס ספק")
    col_invoice_date = headers.get("תאריך חשבונית") or headers.get("ת. חשבונית")

    # Statistics
    stats = ProcessingStats()
    
    # Detailed data storage
    details = ProcessingDetails()

    # Helper function to extract row data
    def extract_row_data(row):
        acc_val = row[col_acc - 1].value
        name_val = row[col_name - 1].value if col_name else ""
        amt_val = row[col_amt - 1].value
        date_val = row[col_pay - 1].value if col_pay else ""
        details_val = row[col_details - 1].value if col_details else ""
        invoice_val = row[col_invoice - 1].value if col_invoice else ""
        acc_desc_val = row[col_acc_desc - 1].value if col_acc_desc else ""
        supplier_acc_val = row[col_supplier_acc - 1].value if col_supplier_acc else ""
        invoice_date_val = row[col_invoice_date - 1].value if col_invoice_date else date_val
        
        # Format date
        if isinstance(date_val, datetime):
            date_str = date_val.strftime("%d/%m/%Y")
        else:
            date_str = str(date_val) if date_val else ""
        
        # Format invoice date
        if isinstance(invoice_date_val, datetime):
            invoice_date_str = invoice_date_val.strftime("%d/%m/%Y")
        else:
            invoice_date_str = str(invoice_date_val) if invoice_date_val else date_str
        
        # Parse amount
        try:
            amount = parse_amount(amt_val)
        except:
            amount = 0
            
        return DetailedRow(
            account=str(acc_val) if acc_val else "",
            account_description=str(acc_desc_val) if acc_desc_val else str(name_val) if name_val else "",
            supplier_account=str(supplier_acc_val) if supplier_acc_val else str(acc_val) if acc_val else "",
            name=str(name_val) if name_val else "",
            amount=amount,
            date=invoice_date_str,
            details=str(details_val) if details_val else "",
            invoice=str(invoice_val) if invoice_val else ""
        )

    # ===== Logic 1 - Green 100% within supplier =====
    groups = defaultdict(list)
    for row in ws.iter_rows(min_row=data_start_row):
        acc = row[col_acc - 1].value
        groups[acc].append(row)

    green_counts = defaultdict(int)

    for acc, rows in groups.items():
        pos, neg = [], []
        for r in rows:
            cell = r[col_amt - 1]
            try:
                v = parse_amount(cell.value)
            except Exception:
                continue
            if v > 0:
                pos.append((v, r))
            elif v < 0:
                neg.append((v, r))

        used_neg = set()
        for pval, prow in pos:
            for ni, (nval, nrow) in enumerate(neg):
                if ni in used_neg:
                    continue
                if abs(pval + nval) < 1e-6:
                    prow[col_amt - 1].fill = GREEN_FILL
                    nrow[col_amt - 1].fill = GREEN_FILL
                    green_counts[acc] += 2
                    # Store detailed data
                    details.green.append(extract_row_data(prow))
                    details.green.append(extract_row_data(nrow))
                    used_neg.add(ni)
                    break

    ensure_summary_sheet(wb, "התאמה 100%", green_counts)
    stats.green_matches = sum(green_counts.values())

    # ===== Logic 3 - Orange 80% within supplier =====
    orange_counts = defaultdict(int)

    for acc, rows in groups.items():
        pos, neg = [], []
        for r in rows:
            cell = r[col_amt - 1]
            if has_any_color(cell):
                continue
            try:
                v = parse_amount(cell.value)
            except Exception:
                continue
            if v > 0:
                pos.append((v, r))
            elif v < 0:
                neg.append((v, r))

        used_neg = set()
        for pval, prow in pos:
            pc = prow[col_amt - 1]
            if has_any_color(pc):
                continue
            for ni, (nval, nrow) in enumerate(neg):
                if ni in used_neg:
                    continue
                nc = nrow[col_amt - 1]
                if has_any_color(nc):
                    continue
                if abs(pval + nval) <= 2:
                    pc.fill = ORANGE_FILL
                    nc.fill = ORANGE_FILL
                    orange_counts[acc] += 2
                    # Store detailed data
                    details.orange.append(extract_row_data(prow))
                    details.orange.append(extract_row_data(nrow))
                    used_neg.add(ni)
                    break

    ensure_summary_sheet(wb, "התאמה 80%", orange_counts)
    stats.orange_matches = sum(orange_counts.values())

    # ===== Logic 5 - Purple global =====
    purple_counts = defaultdict(int)
    eligible = []

    for row in ws.iter_rows(min_row=data_start_row):
        cell = row[col_amt - 1]
        if has_any_color(cell):
            continue
        try:
            v = parse_amount(cell.value)
        except Exception:
            continue
        if v == 0:
            continue
        acc = row[col_acc - 1].value
        eligible.append((v, acc, row))

    pos = [x for x in eligible if x[0] > 0]
    neg = [x for x in eligible if x[0] < 0]

    used_pos, used_neg = set(), set()

    for pi, (pval, pacc, prow) in enumerate(pos):
        if pi in used_pos:
            continue
        pc = prow[col_amt - 1]
        if has_any_color(pc):
            continue
        for ni, (nval, nacc, nrow) in enumerate(neg):
            if ni in used_neg:
                continue
            nc = nrow[col_amt - 1]
            if has_any_color(nc):
                continue
            if abs(pval + nval) <= 2:
                pc.fill = PURPLE_FILL
                nc.fill = PURPLE_FILL
                used_pos.add(pi)
                used_neg.add(ni)
                purple_counts[pacc] += 1
                purple_counts[nacc] += 1
                # Store detailed data
                details.purple.append(extract_row_data(prow))
                details.purple.append(extract_row_data(nrow))
                break

    ensure_summary_sheet(wb, "בדיקת ספקים", purple_counts)
    stats.purple_matches = sum(purple_counts.values())

    # ===== Logic 6 - Blue: transaction type 'העב' + collect for emails =====
    rows_mail = []

    for row in ws.iter_rows(min_row=data_start_row):
        if col_type is None:
            continue
        tval = row[col_type - 1].value
        tval = str(tval).strip() if tval is not None else ""
        cell = row[col_amt - 1]
        if tval == "העב" and not has_any_color(cell):
            cell.fill = BLUE_FILL
            rows_mail.append((
                row[col_name - 1].value,
                row[col_pay - 1].value,
                row[col_amt - 1].value,
                row[col_acc - 1].value,
            ))
            # Store detailed data
            details.blue.append(extract_row_data(row))

    stats.blue_matches = len(rows_mail)

    # ===== Logic 7 - Emails sheet grouped by account =====
    grouped_mail = defaultdict(list)
    for name, pay, debt, acc in rows_mail:
        grouped_mail[str(acc).strip()].append((name, pay, debt))

    if "מיילים לספק" in wb.sheetnames:
        ws_mail = wb["מיילים לספק"]
        for r in ws_mail.iter_rows():
            for c in r:
                c.value = None
    else:
        ws_mail = wb.create_sheet("מיילים לספק")

    ws_mail["A1"] = "שם ספק"
    ws_mail["B1"] = "טקסט מייל"
    ws_mail["C1"] = "מייל ספק"

    row_idx = 2

    for acc, entries in grouped_mail.items():
        name = entries[0][0]
        lines = []
        for _, pay, debt in entries:
            if isinstance(pay, datetime):
                date_str = pay.strftime("%d/%m/%y")
            else:
                date_str = str(pay) if pay else ""
            try:
                amount = abs(parse_amount(debt))
            except Exception:
                amount = debt
            lines.append(f"תאריך - {date_str}\nעל סכום - {amount}")
            
            # Add to emails details
            details.emails.append(DetailedRow(
                account=str(acc) if acc else "",
                name=str(name) if name else "",
                amount=float(amount) if isinstance(amount, (int, float)) else 0,
                date=date_str,
                details="",
                invoice=""
            ))

        combined_details = "\n".join(lines)

        msg = (
            f"שלום ל-{name}\n"
            f"חסרות לנו חשבוניות עבור תשלום:\n"
            f"{combined_details}\n"
            f"בתודה מראש,\n"
            f"הנהלת חשבונות של {company_name}"
        )

        ws_mail.cell(row_idx, 1, name)
        cell_msg = ws_mail.cell(row_idx, 2, msg)
        cell_msg.alignment = Alignment(wrap_text=True)

        supplier_email = ""
        if email_mapping:
            supplier_email = email_mapping.get(acc, "")
            if not supplier_email and name:
                supplier_email = email_mapping.get(str(name).strip(), "")
        if supplier_email:
            ws_mail.cell(row_idx, 3, supplier_email)

        row_idx += 1

    stats.emails_generated = row_idx - 2

    # RTL for all sheets
    for sh in wb.worksheets:
        sh.sheet_view.rightToLeft = True

    # Count total rows
    stats.total_rows = ws.max_row - header_row

    # Calculate special treatment (rows without any color) and save their details
    unmatched_count = 0
    special_rows = []
    for row in ws.iter_rows(min_row=data_start_row):
        cell = row[col_amt - 1]
        if not has_any_color(cell):
            # Check if the cell has a value
            try:
                v = parse_amount(cell.value)
                if v != 0:
                    unmatched_count += 1
                    # Save the row details to special category
                    special_rows.append(extract_row_data(row))
            except Exception:
                pass
    stats.special_treatment = unmatched_count
    
    # Add special rows to details (will be saved to DB)
    details.special = special_rows

    return wb, stats, details


# ========= API Routes =========

@api_router.get("/")
async def root():
    return {"message": "RISE Pro - Debt Aging Automation API"}


@api_router.post("/status", response_model=StatusCheck)
async def create_status_check(input: StatusCheckCreate):
    status_dict = input.model_dump()
    status_obj = StatusCheck(**status_dict)
    doc = status_obj.model_dump()
    doc['timestamp'] = doc['timestamp'].isoformat()
    _ = await db.status_checks.insert_one(doc)
    return status_obj


@api_router.get("/status", response_model=List[StatusCheck])
async def get_status_checks():
    status_checks = await db.status_checks.find({}, {"_id": 0}).to_list(1000)
    for check in status_checks:
        if isinstance(check['timestamp'], str):
            check['timestamp'] = datetime.fromisoformat(check['timestamp'])
    return status_checks


# ========= Supplier Management =========

@api_router.post("/suppliers", response_model=Supplier)
async def create_supplier(supplier: SupplierCreate):
    supplier_obj = Supplier(**supplier.model_dump())
    doc = supplier_obj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    doc['updated_at'] = doc['updated_at'].isoformat()
    await db.suppliers.insert_one(doc)
    return supplier_obj


@api_router.get("/suppliers", response_model=List[Supplier])
async def get_suppliers():
    suppliers = await db.suppliers.find({}, {"_id": 0}).to_list(1000)
    for s in suppliers:
        if isinstance(s.get('created_at'), str):
            s['created_at'] = datetime.fromisoformat(s['created_at'])
        if isinstance(s.get('updated_at'), str):
            s['updated_at'] = datetime.fromisoformat(s['updated_at'])
    return suppliers


@api_router.get("/suppliers/template")
async def download_suppliers_template():
    """Download Excel template for importing suppliers."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ספקים"
    
    # RTL
    ws.sheet_view.rightToLeft = True
    
    # Headers
    headers = ["מס' ספק", "שם ספק", "מטבע", "מס. עוסק מורשה", "חשבון קניות", "תאור חשבון קניות", "טלפון", "e-mail"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = PatternFill(start_color="FF00CDB8", end_color="FF00CDB8", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Example row
    example = ["20001", "שם הספק לדוגמה", "ש\"ח", "123456789", "75001", "קניות דשא / עצים / שתילים", "04-1234567", "example@email.com"]
    for col, val in enumerate(example, 1):
        ws.cell(row=2, column=col, value=val)
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 35
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 25
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=suppliers_template.xlsx"}
    )


@api_router.get("/suppliers/export")
async def export_suppliers_to_excel():
    """Export all suppliers to Excel file."""
    suppliers = await db.suppliers.find({}, {"_id": 0}).to_list(10000)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ספקים"
    ws.sheet_view.rightToLeft = True
    
    # Headers
    headers = ["מס' ספק", "שם ספק", "מטבע", "מס. עוסק מורשה", "חשבון קניות", "תאור חשבון קניות", "טלפון", "e-mail"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = PatternFill(start_color="FF00CDB8", end_color="FF00CDB8", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Data rows
    for row_idx, supplier in enumerate(suppliers, 2):
        ws.cell(row=row_idx, column=1, value=supplier.get("account_number", ""))
        ws.cell(row=row_idx, column=2, value=supplier.get("name", ""))
        ws.cell(row=row_idx, column=3, value=supplier.get("currency", "ש\"ח"))
        ws.cell(row=row_idx, column=4, value=supplier.get("vat_number", ""))
        ws.cell(row=row_idx, column=5, value=supplier.get("purchase_account", ""))
        ws.cell(row=row_idx, column=6, value=supplier.get("purchase_account_desc", ""))
        ws.cell(row=row_idx, column=7, value=supplier.get("phone", ""))
        ws.cell(row=row_idx, column=8, value=supplier.get("email", ""))
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 35
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 25
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=suppliers_export.xlsx"}
    )


@api_router.post("/suppliers/import", response_model=SupplierImportResult)
async def import_suppliers_from_excel(
    file: UploadFile = File(...),
    replace_all: bool = False
):
    """Import suppliers from Excel file."""
    try:
        content = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
        
        # Find header row
        header_row = 1
        headers = {}
        for row_idx in [1, 2]:
            row_cells = ws[row_idx]
            tmp_headers = {str(c.value).strip() if c.value else "": c.column for c in row_cells}
            if "מס' ספק" in tmp_headers or "מס ספק" in tmp_headers:
                header_row = row_idx
                headers = tmp_headers
                break
            if not headers:
                headers = tmp_headers
                header_row = row_idx
        
        # Map column names
        col_account = headers.get("מס' ספק") or headers.get("מס ספק") or headers.get("חשבון")
        col_name = headers.get("שם ספק") or headers.get("שם")
        col_currency = headers.get("מטבע")
        col_vat = headers.get("מס. עוסק מורשה") or headers.get("עוסק מורשה")
        col_purchase = headers.get("חשבון קניות")
        col_purchase_desc = headers.get("תאור חשבון קניות") or headers.get("תיאור חשבון קניות")
        col_phone = headers.get("טלפון")
        col_email = headers.get("e-mail") or headers.get("מייל") or headers.get("Email")
        
        if not col_account or not col_name:
            raise HTTPException(
                status_code=400, 
                detail="הקובץ חייב להכיל עמודות 'מס' ספק' ו'שם ספק'"
            )
        
        result = SupplierImportResult()
        
        # If replace_all, delete existing suppliers
        if replace_all:
            await db.suppliers.delete_many({})
        
        # Process rows
        for row in ws.iter_rows(min_row=header_row + 1):
            result.total += 1
            
            try:
                account_val = row[col_account - 1].value
                name_val = row[col_name - 1].value
                
                if not account_val or not name_val:
                    continue
                
                account_number = str(account_val).strip()
                supplier_name = str(name_val).strip()
                
                # Get optional fields
                currency = str(row[col_currency - 1].value).strip() if col_currency and row[col_currency - 1].value else "ש\"ח"
                vat_number = str(row[col_vat - 1].value).strip() if col_vat and row[col_vat - 1].value else None
                purchase_account = str(row[col_purchase - 1].value).strip() if col_purchase and row[col_purchase - 1].value else None
                purchase_account_desc = str(row[col_purchase_desc - 1].value).strip() if col_purchase_desc and row[col_purchase_desc - 1].value else None
                phone = str(row[col_phone - 1].value).strip() if col_phone and row[col_phone - 1].value else None
                email = str(row[col_email - 1].value).strip() if col_email and row[col_email - 1].value else None
                
                # Clean up phone/email if they are "0" or empty
                if phone in ["0", "None", ""]:
                    phone = None
                if email in ["0", "None", ""]:
                    email = None
                
                # Check if supplier already exists (by account_number)
                existing = await db.suppliers.find_one({"account_number": account_number})
                
                now = datetime.now(timezone.utc).isoformat()
                
                if existing:
                    # Update existing supplier
                    update_data = {
                        "name": supplier_name,
                        "currency": currency,
                        "vat_number": vat_number,
                        "purchase_account": purchase_account,
                        "purchase_account_desc": purchase_account_desc,
                        "phone": phone,
                        "email": email,
                        "updated_at": now
                    }
                    await db.suppliers.update_one(
                        {"account_number": account_number},
                        {"$set": update_data}
                    )
                    result.updated += 1
                else:
                    # Create new supplier
                    supplier_doc = {
                        "id": str(uuid.uuid4()),
                        "account_number": account_number,
                        "name": supplier_name,
                        "currency": currency,
                        "vat_number": vat_number,
                        "purchase_account": purchase_account,
                        "purchase_account_desc": purchase_account_desc,
                        "phone": phone,
                        "email": email,
                        "created_at": now,
                        "updated_at": now
                    }
                    await db.suppliers.insert_one(supplier_doc)
                    result.imported += 1
                    
            except Exception as e:
                result.errors += 1
                result.error_messages.append(f"שורה {result.total}: {str(e)}")
        
        return result
        
    except Exception as e:
        logger.error(f"Import error: {e}")
        raise HTTPException(status_code=500, detail=f"שגיאה בייבוא: {str(e)}")


@api_router.get("/suppliers/{supplier_id}", response_model=Supplier)
async def get_supplier(supplier_id: str):
    supplier = await db.suppliers.find_one({"id": supplier_id}, {"_id": 0})
    if not supplier:
        raise HTTPException(status_code=404, detail="Supplier not found")
    if isinstance(supplier.get('created_at'), str):
        supplier['created_at'] = datetime.fromisoformat(supplier['created_at'])
    if isinstance(supplier.get('updated_at'), str):
        supplier['updated_at'] = datetime.fromisoformat(supplier['updated_at'])
    return supplier


@api_router.put("/suppliers/{supplier_id}", response_model=Supplier)
async def update_supplier(supplier_id: str, update: SupplierUpdate):
    existing = await db.suppliers.find_one({"id": supplier_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Supplier not found")
    
    update_data = {k: v for k, v in update.model_dump().items() if v is not None}
    update_data['updated_at'] = datetime.now(timezone.utc).isoformat()
    
    await db.suppliers.update_one({"id": supplier_id}, {"$set": update_data})
    return await get_supplier(supplier_id)


@api_router.delete("/suppliers/{supplier_id}")
async def delete_supplier(supplier_id: str):
    result = await db.suppliers.delete_one({"id": supplier_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Supplier not found")
    return {"message": "Supplier deleted successfully"}


@api_router.delete("/suppliers")
async def delete_all_suppliers():
    """Delete all suppliers from the database."""
    result = await db.suppliers.delete_many({})
    return {"message": f"Deleted {result.deleted_count} suppliers", "count": result.deleted_count}




# ========= Processing History =========

@api_router.get("/processing-history", response_model=List[ProcessingHistory])
async def get_processing_history():
    history = await db.processing_history.find({}, {"_id": 0}).sort("processed_at", -1).to_list(100)
    for h in history:
        if isinstance(h.get('processed_at'), str):
            h['processed_at'] = datetime.fromisoformat(h['processed_at'])
    return history


# ========= N8N Trigger =========

@api_router.post("/trigger-n8n")
async def trigger_n8n(request: N8NTriggerRequest):
    if not N8N_WEBHOOK_URL:
        raise HTTPException(status_code=500, detail="N8N Webhook URL not configured")
    
    payload = {
        "client_name": request.client_name,
        "action": "giyul_chovot",
    }
    
    try:
        resp = requests.post(N8N_WEBHOOK_URL, json=payload, timeout=15)
        resp.raise_for_status()
        return {"success": True, "message": f"Trigger sent for: {request.client_name}"}
    except Exception as e:
        logger.error(f"N8N trigger failed: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to trigger N8N: {str(e)}")


# ========= File Processing =========

@api_router.post("/process-excel")
async def process_excel(
    main_file: UploadFile = File(...),
    helper_file: Optional[UploadFile] = File(None)
):
    """Process debt aging Excel file with all logics."""
    try:
        # Read main file
        main_content = await main_file.read()
        wb = openpyxl.load_workbook(io.BytesIO(main_content))
        
        # Build email mapping if helper file provided
        email_mapping = None
        if helper_file:
            helper_content = await helper_file.read()
            helper_wb = openpyxl.load_workbook(io.BytesIO(helper_content), data_only=True)
            email_mapping = build_email_mapping(helper_wb)
        
        # Process workbook
        wb, stats, details = process_workbook(wb, email_mapping=email_mapping)
        
        # Save details to database for later retrieval
        details_doc = {
            "id": str(uuid.uuid4()),
            "filename": main_file.filename or "unknown.xlsx",
            "processed_at": datetime.now(timezone.utc).isoformat(),
            "green": [d.model_dump() for d in details.green],
            "orange": [d.model_dump() for d in details.orange],
            "purple": [d.model_dump() for d in details.purple],
            "blue": [d.model_dump() for d in details.blue],
            "emails": [d.model_dump() for d in details.emails],
            "special": [d.model_dump() for d in details.special],  # Rows without any match
            "command": [],  # Will be populated when user moves rows here
        }
        await db.processing_details.delete_many({})  # Keep only latest
        await db.processing_details.insert_one(details_doc)
        
        # Save processing history
        history = ProcessingHistory(
            filename=main_file.filename or "unknown.xlsx",
            green_matches=stats.green_matches,
            orange_matches=stats.orange_matches,
            purple_matches=stats.purple_matches,
            blue_matches=stats.blue_matches,
            total_rows=stats.total_rows
        )
        history_doc = history.model_dump()
        history_doc['processed_at'] = history_doc['processed_at'].isoformat()
        await db.processing_history.insert_one(history_doc)
        
        # Save to buffer
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Return file - use URL encoding for Hebrew filename
        safe_filename = quote(main_file.filename or "output.xlsx")
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename*=UTF-8''{safe_filename}",
                "X-Stats-Green": str(stats.green_matches),
                "X-Stats-Orange": str(stats.orange_matches),
                "X-Stats-Purple": str(stats.purple_matches),
                "X-Stats-Blue": str(stats.blue_matches),
                "X-Stats-Special": str(stats.special_treatment),
                "X-Stats-Total": str(stats.total_rows),
                "X-Stats-Emails": str(stats.emails_generated),
                "Access-Control-Expose-Headers": "X-Stats-Green,X-Stats-Orange,X-Stats-Purple,X-Stats-Blue,X-Stats-Special,X-Stats-Total,X-Stats-Emails,Content-Disposition"
            }
        )
        
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        logger.error(f"Processing error: {e}")
        raise HTTPException(status_code=500, detail=f"Error processing file: {str(e)}")


@api_router.get("/processing-details/{category}")
async def get_processing_details(category: str):
    """Get detailed rows for a specific category (green, orange, purple, blue, special, command, emails)."""
    if category not in ["green", "orange", "purple", "blue", "special", "command", "emails"]:
        raise HTTPException(status_code=400, detail="Invalid category")
    
    details = await db.processing_details.find_one({}, {"_id": 0})
    if not details:
        return {"rows": []}
    
    return {"rows": details.get(category, [])}


@api_router.post("/move-row")
async def move_row(request: MoveRowRequest):
    """Move a row from one category to another."""
    valid_categories = ["green", "orange", "purple", "blue", "special", "command", "emails"]
    if request.from_category not in valid_categories or request.to_category not in valid_categories:
        raise HTTPException(status_code=400, detail="Invalid category")
    
    details = await db.processing_details.find_one({}, {"_id": 0})
    if not details:
        raise HTTPException(status_code=404, detail="No processing data found")
    
    # Get source and target lists
    source_list = details.get(request.from_category, [])
    target_list = details.get(request.to_category, [])
    
    # Find and remove from source by matching row data
    row_to_move = None
    for i, row in enumerate(source_list):
        if (row.get("account") == request.row_data.get("account") and 
            row.get("name") == request.row_data.get("name") and
            abs(row.get("amount", 0) - request.row_data.get("amount", 0)) < 0.01 and
            row.get("date") == request.row_data.get("date")):
            row_to_move = source_list.pop(i)
            break
    
    if not row_to_move:
        raise HTTPException(status_code=404, detail="Row not found in source category")
    
    # Add to target
    target_list.append(row_to_move)
    
    # Update database
    await db.processing_details.update_one(
        {},
        {"$set": {
            request.from_category: source_list,
            request.to_category: target_list
        }}
    )
    
    return {"success": True, "message": f"Row moved from {request.from_category} to {request.to_category}"}


class DeleteRowRequest(BaseModel):
    from_category: str
    row_data: dict


@api_router.post("/delete-row")
async def delete_row(request: DeleteRowRequest):
    """Delete a row from a category (remove completely without moving)."""
    valid_categories = ["green", "orange", "purple", "blue", "special", "command", "emails"]
    if request.from_category not in valid_categories:
        raise HTTPException(status_code=400, detail="Invalid category")
    
    details = await db.processing_details.find_one({}, {"_id": 0})
    if not details:
        raise HTTPException(status_code=404, detail="No processing data found")
    
    # Get source list
    source_list = details.get(request.from_category, [])
    
    # Find and remove from source by matching row data
    row_found = False
    for i, row in enumerate(source_list):
        if (row.get("account") == request.row_data.get("account") and 
            row.get("name") == request.row_data.get("name") and
            abs(row.get("amount", 0) - request.row_data.get("amount", 0)) < 0.01 and
            row.get("date") == request.row_data.get("date")):
            source_list.pop(i)
            row_found = True
            break
    
    if not row_found:
        raise HTTPException(status_code=404, detail="Row not found in category")
    
    # Update database
    await db.processing_details.update_one(
        {},
        {"$set": {request.from_category: source_list}}
    )
    
    return {"success": True, "message": f"Row deleted from {request.from_category}"}


class DeleteSupplierRowsRequest(BaseModel):
    category: str
    supplier_name: str
    supplier_account: str


@api_router.post("/delete-supplier-rows")
async def delete_supplier_rows(request: DeleteSupplierRowsRequest):
    """Delete all rows for a specific supplier from a category."""
    valid_categories = ["green", "orange", "purple", "blue", "special", "command", "emails"]
    if request.category not in valid_categories:
        raise HTTPException(status_code=400, detail="Invalid category")
    
    details = await db.processing_details.find_one({}, {"_id": 0})
    if not details:
        raise HTTPException(status_code=404, detail="No processing data found")
    
    # Get source list
    source_list = details.get(request.category, [])
    
    # Count rows before deletion
    original_count = len(source_list)
    
    # Filter out all rows matching the supplier
    source_list = [row for row in source_list if not (
        row.get("name") == request.supplier_name or 
        row.get("account") == request.supplier_account
    )]
    
    deleted_count = original_count - len(source_list)
    
    if deleted_count == 0:
        raise HTTPException(status_code=404, detail="No rows found for this supplier")
    
    # Update database
    await db.processing_details.update_one(
        {},
        {"$set": {request.category: source_list}}
    )
    
    return {"success": True, "deleted_count": deleted_count, "message": f"Deleted {deleted_count} rows for supplier"}


class PaymentRequest(BaseModel):
    rows: List[dict]
    supplier_name: str

@api_router.post("/generate-payment")
async def generate_payment(request: PaymentRequest):
    """Generate Excel payment file for selected rows."""
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "רשימת תשלום"
        ws.sheet_view.rightToLeft = True
        
        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Headers
        headers = ["מס'", "חשבון", "שם ספק", "תאור חשבון", "חש. ספק", "סכום", "תאריך חשבונית", "תאריך תשלום", "פרטים"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Calculate payment date (invoice date + 60 days, then to the 10th of that month)
        def calculate_payment_date(invoice_date_str):
            try:
                # Try to parse the date
                if "/" in str(invoice_date_str):
                    parts = str(invoice_date_str).split("/")
                    if len(parts) == 3:
                        day, month, year = int(parts[0]), int(parts[1]), int(parts[2])
                        if year < 100:
                            year += 2000
                        invoice_date = datetime(year, month, day)
                    else:
                        invoice_date = datetime.now()
                else:
                    invoice_date = datetime.now()
                
                # Add 60 days
                payment_date = invoice_date + timedelta(days=60)
                # Set to 10th of that month
                payment_date = payment_date.replace(day=10)
                return payment_date.strftime("%d/%m/%Y")
            except:
                return ""
        
        # Data rows
        total_amount = 0
        for idx, row in enumerate(request.rows, 1):
            ws.cell(row=idx+1, column=1, value=idx)
            ws.cell(row=idx+1, column=2, value=row.get("account", ""))
            ws.cell(row=idx+1, column=3, value=row.get("name", ""))
            ws.cell(row=idx+1, column=4, value=row.get("account_description", ""))
            ws.cell(row=idx+1, column=5, value=row.get("supplier_account", ""))
            
            amount = row.get("amount", 0)
            total_amount += amount
            amount_cell = ws.cell(row=idx+1, column=6, value=amount)
            amount_cell.number_format = '#,##0.00 ₪'
            
            ws.cell(row=idx+1, column=7, value=row.get("date", ""))
            ws.cell(row=idx+1, column=8, value=calculate_payment_date(row.get("date", "")))
            ws.cell(row=idx+1, column=9, value=row.get("details", ""))
        
        # Total row
        total_row = len(request.rows) + 2
        ws.cell(row=total_row, column=5, value="סה״כ:").font = Font(bold=True)
        total_cell = ws.cell(row=total_row, column=6, value=total_amount)
        total_cell.font = Font(bold=True)
        total_cell.number_format = '#,##0.00 ₪'
        
        # Auto-fit columns
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column].width = max_length + 2
        
        # Remove rows from special category in DB
        details = await db.processing_details.find_one({}, {"_id": 0})
        if details:
            special_list = details.get("special", [])
            row_keys = set(f"{r.get('account')}-{r.get('amount')}-{r.get('date')}" for r in request.rows)
            special_list = [r for r in special_list if f"{r.get('account')}-{r.get('amount')}-{r.get('date')}" not in row_keys]
            await db.processing_details.update_one({}, {"$set": {"special": special_list}})
        
        # Save to buffer
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        safe_filename = quote(f"תשלום_{request.supplier_name}.xlsx")
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename*=UTF-8''{safe_filename}"
            }
        )
    except Exception as e:
        logger.error(f"Error generating payment: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/send-email")
async def send_email(request: SendEmailRequest):
    """Send email via SMTP (Outlook/Office 365)."""
    try:
        # Create message
        msg = MIMEMultipart('alternative')
        msg['Subject'] = request.subject
        msg['From'] = f"{request.sender_name} <{request.sender_email}>" if request.sender_name else request.sender_email
        msg['To'] = request.recipient_email
        
        # Create plain text and HTML versions
        text_body = request.body
        html_body = f"""
        <html dir="rtl">
        <body style="font-family: Arial, sans-serif; direction: rtl; text-align: right;">
            {request.body.replace(chr(10), '<br>')}
        </body>
        </html>
        """
        
        part1 = MIMEText(text_body, 'plain', 'utf-8')
        part2 = MIMEText(html_body, 'html', 'utf-8')
        msg.attach(part1)
        msg.attach(part2)
        
        # Connect to Outlook SMTP server
        with smtplib.SMTP('smtp.office365.com', 587) as server:
            server.starttls()
            server.login(request.sender_email, request.sender_password)
            server.sendmail(request.sender_email, request.recipient_email, msg.as_string())
        
        logger.info(f"Email sent successfully from {request.sender_email} to {request.recipient_email}")
        return {"success": True, "message": "המייל נשלח בהצלחה"}
    
    except smtplib.SMTPAuthenticationError as e:
        logger.error(f"SMTP Authentication failed: {e}")
        raise HTTPException(status_code=401, detail="שגיאת אימות - בדקי את המייל וסיסמת האפליקציה")
    except smtplib.SMTPException as e:
        logger.error(f"SMTP Error: {e}")
        raise HTTPException(status_code=500, detail=f"שגיאה בשליחת המייל: {str(e)}")
    except Exception as e:
        logger.error(f"Email sending error: {e}")
        raise HTTPException(status_code=500, detail=f"שגיאה: {str(e)}")


@api_router.post("/preview-excel")
async def preview_excel(file: UploadFile = File(...)):
    """Preview first few rows of uploaded Excel file."""
    try:
        content = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
        
        header_row, headers = detect_headers(ws)
        
        # Get first 10 data rows
        preview_data = []
        row_count = 0
        for row in ws.iter_rows(min_row=header_row + 1, max_row=header_row + 11):
            row_data = {}
            for header, col_idx in headers.items():
                cell_value = row[col_idx - 1].value
                if isinstance(cell_value, datetime):
                    cell_value = cell_value.strftime("%d/%m/%Y")
                row_data[header] = cell_value
            preview_data.append(row_data)
            row_count += 1
        
        total_rows = ws.max_row - header_row
        
        return {
            "headers": list(headers.keys()),
            "preview": preview_data,
            "total_rows": total_rows,
            "header_row": header_row
        }
        
    except Exception as e:
        logger.error(f"Preview error: {e}")
        raise HTTPException(status_code=500, detail=f"Error previewing file: {str(e)}")


# ========= Microsoft OAuth2 Endpoints =========

@api_router.get("/auth/microsoft/login")
async def microsoft_login():
    """Initiate Microsoft OAuth2 login flow."""
    scope = " ".join(MICROSOFT_SCOPES)
    auth_url = (
        f"{MICROSOFT_AUTHORITY}/oauth2/v2.0/authorize"
        f"?client_id={MICROSOFT_CLIENT_ID}"
        f"&response_type=code"
        f"&redirect_uri={quote(MICROSOFT_REDIRECT_URI)}"
        f"&response_mode=query"
        f"&scope={quote(scope)}"
        f"&prompt=select_account"
    )
    return {"auth_url": auth_url}


@api_router.get("/auth/microsoft/callback")
async def microsoft_callback(code: str = None, error: str = None, error_description: str = None):
    """Handle Microsoft OAuth2 callback."""
    if error:
        logger.error(f"OAuth error: {error} - {error_description}")
        return RedirectResponse(url=f"{FRONTEND_URL}?auth_error={error}")
    
    if not code:
        return RedirectResponse(url=f"{FRONTEND_URL}?auth_error=no_code")
    
    try:
        # Exchange code for tokens
        token_url = f"{MICROSOFT_AUTHORITY}/oauth2/v2.0/token"
        token_data = {
            "client_id": MICROSOFT_CLIENT_ID,
            "client_secret": MICROSOFT_CLIENT_SECRET,
            "code": code,
            "redirect_uri": MICROSOFT_REDIRECT_URI,
            "grant_type": "authorization_code",
            "scope": " ".join(MICROSOFT_SCOPES)
        }
        
        async with httpx.AsyncClient() as client_http:
            token_response = await client_http.post(token_url, data=token_data)
            
            if token_response.status_code != 200:
                logger.error(f"Token exchange failed: {token_response.text}")
                return RedirectResponse(url=f"{FRONTEND_URL}?auth_error=token_exchange_failed")
            
            tokens = token_response.json()
            access_token = tokens.get("access_token")
            refresh_token = tokens.get("refresh_token")
            
            # Get user info
            user_response = await client_http.get(
                "https://graph.microsoft.com/v1.0/me",
                headers={"Authorization": f"Bearer {access_token}"}
            )
            
            if user_response.status_code != 200:
                logger.error(f"User info failed: {user_response.text}")
                return RedirectResponse(url=f"{FRONTEND_URL}?auth_error=user_info_failed")
            
            user_info = user_response.json()
            
            # Save user tokens to database
            user_data = {
                "microsoft_id": user_info.get("id"),
                "email": user_info.get("mail") or user_info.get("userPrincipalName"),
                "display_name": user_info.get("displayName"),
                "access_token": access_token,
                "refresh_token": refresh_token,
                "updated_at": datetime.now(timezone.utc).isoformat()
            }
            
            await db.microsoft_users.update_one(
                {"email": user_data["email"]},
                {"$set": user_data},
                upsert=True
            )
            
            logger.info(f"Microsoft user logged in: {user_data['email']}")
            
            # Redirect back to frontend with success
            return RedirectResponse(url=f"{FRONTEND_URL}?auth_success=true&email={quote(user_data['email'])}&name={quote(user_data['display_name'] or '')}")
    
    except Exception as e:
        logger.error(f"OAuth callback error: {e}")
        return RedirectResponse(url=f"{FRONTEND_URL}?auth_error=server_error")


@api_router.get("/auth/microsoft/status")
async def microsoft_auth_status(email: str):
    """Check if user is authenticated with Microsoft."""
    user = await db.microsoft_users.find_one({"email": email}, {"_id": 0})
    if user:
        return {
            "authenticated": True,
            "email": user.get("email"),
            "display_name": user.get("display_name")
        }
    return {"authenticated": False}


@api_router.post("/auth/microsoft/logout")
async def microsoft_logout(email: str):
    """Logout user from Microsoft."""
    await db.microsoft_users.delete_one({"email": email})
    return {"success": True}


async def refresh_microsoft_token(user_email: str):
    """Refresh Microsoft access token."""
    user = await db.microsoft_users.find_one({"email": user_email})
    if not user or not user.get("refresh_token"):
        return None
    
    try:
        token_url = f"{MICROSOFT_AUTHORITY}/oauth2/v2.0/token"
        token_data = {
            "client_id": MICROSOFT_CLIENT_ID,
            "client_secret": MICROSOFT_CLIENT_SECRET,
            "refresh_token": user["refresh_token"],
            "grant_type": "refresh_token",
            "scope": " ".join(MICROSOFT_SCOPES)
        }
        
        async with httpx.AsyncClient() as client_http:
            response = await client_http.post(token_url, data=token_data)
            
            if response.status_code != 200:
                logger.error(f"Token refresh failed: {response.text}")
                return None
            
            tokens = response.json()
            
            # Update tokens in database
            await db.microsoft_users.update_one(
                {"email": user_email},
                {"$set": {
                    "access_token": tokens.get("access_token"),
                    "refresh_token": tokens.get("refresh_token", user["refresh_token"]),
                    "updated_at": datetime.now(timezone.utc).isoformat()
                }}
            )
            
            return tokens.get("access_token")
    
    except Exception as e:
        logger.error(f"Token refresh error: {e}")
        return None


class MicrosoftEmailRequest(BaseModel):
    sender_email: str
    recipient_email: str
    subject: str
    body: str


@api_router.post("/send-email-microsoft")
async def send_email_microsoft(request: MicrosoftEmailRequest):
    """Send email via Microsoft Graph API."""
    # Get user from database
    user = await db.microsoft_users.find_one({"email": request.sender_email})
    
    if not user:
        raise HTTPException(status_code=401, detail="המשתמש לא מחובר ל-Microsoft. יש להתחבר קודם.")
    
    access_token = user.get("access_token")
    
    # Convert text to HTML with formatting
    import re
    html_body = request.body
    # Convert *** text *** to bold
    html_body = re.sub(r'\*\*\*(.+?)\*\*\*', r'<strong style="color: #00CDB8; font-weight: bold; font-size: 16px;">\1</strong>', html_body)
    # Convert newlines to <br>
    html_body = html_body.replace('\n', '<br>')
    
    # Try to send email
    async with httpx.AsyncClient() as client_http:
        email_data = {
            "message": {
                "subject": request.subject,
                "body": {
                    "contentType": "HTML",
                    "content": f'<div dir="rtl" style="font-family: Arial, sans-serif;">{html_body}</div>'
                },
                "toRecipients": [
                    {"emailAddress": {"address": request.recipient_email}}
                ]
            },
            "saveToSentItems": True
        }
        
        response = await client_http.post(
            "https://graph.microsoft.com/v1.0/me/sendMail",
            json=email_data,
            headers={
                "Authorization": f"Bearer {access_token}",
                "Content-Type": "application/json"
            }
        )
        
        # If token expired, try to refresh and retry
        if response.status_code == 401:
            new_token = await refresh_microsoft_token(request.sender_email)
            if new_token:
                response = await client_http.post(
                    "https://graph.microsoft.com/v1.0/me/sendMail",
                    json=email_data,
                    headers={
                        "Authorization": f"Bearer {new_token}",
                        "Content-Type": "application/json"
                    }
                )
        
        if response.status_code == 202:
            logger.info(f"Email sent via Microsoft Graph from {request.sender_email} to {request.recipient_email}")
            return {"success": True, "message": "המייל נשלח בהצלחה!"}
        else:
            logger.error(f"Microsoft Graph email error: {response.status_code} - {response.text}")
            raise HTTPException(status_code=500, detail=f"שגיאה בשליחת המייל: {response.text}")


# Include router and middleware
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
    expose_headers=["X-Stats-Green", "X-Stats-Orange", "X-Stats-Purple", "X-Stats-Blue", "X-Stats-Total", "X-Stats-Emails", "X-Stats-Special", "X-Stats-Command", "Content-Disposition"]
)


@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()

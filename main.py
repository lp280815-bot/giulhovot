import io
from collections import defaultdict
from datetime import datetime

import openpyxl
from openpyxl.styles import PatternFill, Alignment
from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.responses import StreamingResponse


app = FastAPI()


# ================= Health-check =================

@app.get("/")
async def healthcheck():
    """
    Простой health-check, чтобы видеть что сервис жив.
    """
    return {
        "status": "ok",
        "service": "giulhovot",
        "message": "service is alive",
    }


# ================= ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ =================

def parse_amount(val):
    """Преобразование значения в float (с учётом пустых и разделителей тысяч)."""
    if val is None or val == "":
        raise ValueError("empty")
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    if s == "":
        raise ValueError("empty")
    s = s.replace(",", "")  # убрать разделители тысяч
    return float(s)


def detect_headers(ws):
    """
    Поиск строки заголовков: сначала пытаемся строку 1, потом 2.
    Возвращает: (index строки заголовков, dict {имя колонки -> индекс колонки})
    """
    candidates = [1, 2]
    chosen_row = None
    headers = {}

    for row_idx in candidates:
        row_cells = ws[row_idx]
        row_values = [str(c.value).strip() if c.value is not None else "" for c in row_cells]
        if any(v for v in row_values):
            tmp_headers = {str(c.value).strip(): c.column for c in ws[row_idx] if c.value}
            # приоритет строке, где есть и "חשבון" и "חוב לחשבונית"
            if "חשבון" in tmp_headers and "חוב לחשבונית" in tmp_headers:
                chosen_row = row_idx
                headers = tmp_headers
                break
            # если ещё не выбрали — держим как временную
            if not headers:
                chosen_row = row_idx
                headers = tmp_headers

    # fallback — первая строка
    if chosen_row is None:
        chosen_row = 1
        headers = {str(c.value).strip(): c.column for c in ws[1] if c.value}

    return chosen_row, headers


# ---------- Цвета ----------

GREEN_RGB = "FF00FF00"   # зелёный
ORANGE_RGB = "FFFFA500"  # оранжевый
PURPLE_RGB = "FFCC99FF"  # фиолетовый
BLUE_RGB = "FFADD8E6"    # голубой

GREEN_FILL = PatternFill(start_color=GREEN_RGB, end_color=GREEN_RGB, fill_type="solid")
ORANGE_FILL = PatternFill(start_color=ORANGE_RGB, end_color=ORANGE_RGB, fill_type="solid")
PURPLE_FILL = PatternFill(start_color=PURPLE_RGB, end_color=PURPLE_RGB, fill_type="solid")
BLUE_FILL = PatternFill(start_color=BLUE_RGB, end_color=BLUE_RGB, fill_type="solid")


def cell_rgb(cell):
    try:
        return cell.fill.start_color.rgb
    except Exception:
        return None


def has_any_color(cell):
    """Проверка, есть ли у ячейки один из цветов логик."""
    return cell.fill.fill_type == "solid" and cell_rgb(cell) in {
        GREEN_RGB,
        ORANGE_RGB,
        PURPLE_RGB,
        BLUE_RGB,
    }


# ---------- Лист "סיכום" ----------

def ensure_summary_sheet(wb, title, counts):
    """Создать / очистить лист суммарной статистики и заполнить данными."""
    if title in wb.sheetnames:
        ws_sum = wb[title]
        for row in ws_sum.iter_rows():
            for c in row:
                c.value = None
    else:
        ws_sum = wb.create_sheet(title)

    ws_sum["A1"] = "מס ספק"
    ws_sum["B1"] = "כמות שורות מותאמות"

    r = 2
    for acc, cnt in counts.items():
        if acc is None or cnt <= 0:
            continue
        ws_sum.cell(r, 1, acc)
        ws_sum.cell(r, 2, cnt)
        r += 1


# ---------- Построение словаря e-mail ----------

def build_email_mapping(helper_file):
    """
    Строит словарь {חשבון/שם ספק -> email} из файла-помощника.
    Ищет колонки:
    - 'חשבון' / 'מס ספק'
    - 'שם ספק' / 'תאור חשבון' / 'תיאור חשבון'
    - 'מייל' / 'מייל ספק' / 'Email' / 'E-mail'
    """
    wb_help = openpyxl.load_workbook(helper_file, data_only=True)
    ws_help = wb_help.active

    header_row, headers = detect_headers(ws_help)

    col_acc = headers.get("חשבון") or headers.get("מס ספק")
    col_name = headers.get("שם ספק") or headers.get("תאור חשבון") or headers.get("תיאור חשבון")
    col_email = (
        headers.get("מייל")
        or headers.get("מייל ספק")
        or headers.get("Email")
        or headers.get("E-mail")
    )

    email_map = {}

    if col_email is None:
        return email_map

    # по счёту
    if col_acc is not None:
        for row in ws_help.iter_rows(min_row=header_row + 1):
            acc = row[col_acc - 1].value
            email = row[col_email - 1].value
            if acc and email:
                email_map[str(acc).strip()] = str(email).strip()

    # по имени поставщика
    if col_name is not None:
        for row in ws_help.iter_rows(min_row=header_row + 1):
            name = row[col_name - 1].value
            email = row[col_email - 1].value
            if name and email:
                email_map[str(name).strip()] = str(email).strip()

    return email_map


# ---------- Логики 1–7 ----------

def process_workbook(wb, email_mapping=None):
    """
    Запускает на Workbook все логики 1–7.
    email_mapping – опциональный словарь {חשבון/שם ספק -> email}.
    """
    ws = wb.active  # первый лист — исходный

    header_row, headers = detect_headers(ws)

    col_acc = headers.get("חשבון")          # מס ספק
    col_amt = headers.get("חוב לחשבונית")   # сумма к оплате
    col_type = headers.get("סוג תנועה")     # тип движения
    col_name = headers.get("תאור חשבון") or headers.get("שם ספק") or headers.get("תיאור חשבון")
    col_pay = headers.get("תאריך תשלום")    # дата платежа

    if col_acc is None or col_amt is None:
        raise ValueError("לא נמצאו עמודות 'חשבון' ו/או 'חוב לחשבונית'.")

    if col_name is None:
        col_name = 3
    if col_pay is None:
        col_pay = 4

    data_start_row = header_row + 1

    # название компании – для текста письма
    company_name = ws["C1"].value if ws["C1"].value is not None else ""

    # ===== Логика 1 – зелёный: 100% внутри поставщика =====
    groups = defaultdict(list)
    for row in ws.iter_rows(min_row=data_start_row):
        acc = row[col_acc - 1].value
        groups[acc].append(row)

    green_counts = defaultdict(int)

    for acc, rows in groups.items():
        pos, neg = [], []
        for r in rows:
            cell = r[col_amt - 1]
            try:
                v = parse_amount(cell.value)
            except Exception:
                continue
            if v > 0:
                pos.append((v, r))
            elif v < 0:
                neg.append((v, r))

        used_neg = set()
        for pval, prow in pos:
            for ni, (nval, nrow) in enumerate(neg):
                if ni in used_neg:
                    continue
                if abs(pval + nval) < 1e-6:
                    prow[col_amt - 1].fill = GREEN_FILL
                    nrow[col_amt - 1].fill = GREEN_FILL
                    green_counts[acc] += 2
                    used_neg.add(ni)
                    break

    ensure_summary_sheet(wb, "התאמה 100%", green_counts)

    # ===== Логика 3 – оранжевый: 80% внутри поставщика =====
    orange_counts = defaultdict(int)

    for acc, rows in groups.items():
        pos, neg = [], []
        for r in rows:
            cell = r[col_amt - 1]
            if has_any_color(cell):
                continue
            try:
                v = parse_amount(cell.value)
            except Exception:
                continue
            if v > 0:
                pos.append((v, r))
            elif v < 0:
                neg.append((v, r))

        used_neg = set()
        for pval, prow in pos:
            pc = prow[col_amt - 1]
            if has_any_color(pc):
                continue
            for ni, (nval, nrow) in enumerate(neg):
                if ni in used_neg:
                    continue
                nc = nrow[col_amt - 1]
                if has_any_color(nc):
                    continue
                if abs(pval + nval) <= 2:
                    pc.fill = ORANGE_FILL
                    nc.fill = ORANGE_FILL
                    orange_counts[acc] += 2
                    used_neg.add(ni)
                    break

    ensure_summary_sheet(wb, "התאמה 80%", orange_counts)

    # ===== Логика 5 – фиолетовый: глобальный =====
    purple_counts = defaultdict(int)
    eligible = []

    for row in ws.iter_rows(min_row=data_start_row):
        cell = row[col_amt - 1]
        if has_any_color(cell):
            continue
        try:
            v = parse_amount(cell.value)
        except Exception:
            continue
        if v == 0:
            continue
        acc = row[col_acc - 1].value
        eligible.append((v, acc, row))

    pos = [x for x in eligible if x[0] > 0]
    neg = [x for x in eligible if x[0] < 0]

    used_pos, used_neg = set(), set()

    for pi, (pval, pacc, prow) in enumerate(pos):
        if pi in used_pos:
            continue
        pc = prow[col_amt - 1]
        if has_any_color(pc):
            continue
        for ni, (nval, nacc, nrow) in enumerate(neg):
            if ni in used_neg:
                continue
            nc = nrow[col_amt - 1]
            if has_any_color(nc):
                continue
            if abs(pval + nval) <= 2:
                pc.fill = PURPLE_FILL
                nc.fill = PURPLE_FILL
                used_pos.add(pi)
                used_neg.add(ni)
                purple_counts[pacc] += 1
                purple_counts[nacc] += 1
                break

    ensure_summary_sheet(wb, "בדיקת ספקים", purple_counts)

    # ===== Логика 6 – голубой: סוג תנועה 'העב' + сбор строк для писем =====
    rows_mail = []

    for row in ws.iter_rows(min_row=data_start_row):
        if col_type is None:
            continue
        tval = row[col_type - 1].value
        tval = str(tval).strip() if tval is not None else ""
        cell = row[col_amt - 1]
        if tval == "העב" and not has_any_color(cell):
            cell.fill = BLUE_FILL
            rows_mail.append(
                (
                    row[col_name - 1].value,   # שם ספק
                    row[col_pay - 1].value,    # תאריך תשלום
                    row[col_amt - 1].value,    # חוב לחשבונית
                    row[col_acc - 1].value,    # חשבון
                )
            )

    # ===== Логика 7 – лист "מיילים לספק" сгруппированный по חשבון =====

    grouped_mail = defaultdict(list)
    for name, pay, debt, acc in rows_mail:
        grouped_mail[str(acc).strip()].append((name, pay, debt))

    if "מיילים לספק" in wb.sheetnames:
        ws_mail = wb["מיילים לספק"]
        for r in ws_mail.iter_rows():
            for c in r:
                c.value = None
    else:
        ws_mail = wb.create_sheet("מיילים לספק")

    ws_mail["A1"] = "שם ספק"
    ws_mail["B1"] = "טקסט מייל"
    ws_mail["C1"] = "מייל ספק"

    row_idx = 2

    for acc, entries in grouped_mail.items():
        name = entries[0][0]

        lines = []
        for _, pay, debt in entries:
            if isinstance(pay, datetime):
                date_str = pay.strftime("%d/%m/%y")
            else:
                date_str = str(pay) if pay else ""
            try:
                amount = abs(parse_amount(debt))
            except Exception:
                amount = debt
            lines.append(f"תאריך - {date_str}\nעל סכום - {amount}")

        combined_details = "\n".join(lines)

        msg = (
            f"שלום ל-{name}\n"
            f"חסרות לנו חשבוניות עבור תשלום:\n"
            f"{combined_details}\n"
            f"בתודה מראש,\n"
            f"הנהלת חשבונות של {company_name}"
        )

        ws_mail.cell(row_idx, 1, name)
        cell_msg = ws_mail.cell(row_idx, 2, msg)
        cell_msg.alignment = Alignment(wrap_text=True)

        supplier_email = ""
        if email_mapping:
            supplier_email = email_mapping.get(acc, "")
            if not supplier_email and name:
                supplier_email = email_mapping.get(str(name).strip(), "")
        if supplier_email:
            ws_mail.cell(row_idx, 3, supplier_email)

        row_idx += 1

    # RTL для всех листов
    for sh in wb.worksheets:
        sh.sheet_view.rightToLeft = True

    return wb


# ================= ENDPOINT /process ДЛЯ n8n =================

@app.post("/process")
async def process(
    file1: UploadFile = File(...),          # גיול חובות שוקי.N8N.xlsx
    file2: UploadFile | None = File(None),  # גיול חובות שוקי.N8N.מיילים.xlsx (опционально)
):
    """
    Получает из n8n два файла:
    - file1: гиюль ховот
    - file2: файл-помощник с e-mail поставщиков (может быть None)

    Возвращает готовый Excel с логиками 1–7 и листом "מיילים לספק".
    """
    try:
        # читаем основной файл гиюль
        giyul_bytes = io.BytesIO(await file1.read())
        wb = openpyxl.load_workbook(giyul_bytes, data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"שגיאה בקריאת קובץ גיול חובות: {e}")

    # строим mapping e-mail если есть второй файл
    email_mapping = None
    if file2 is not None:
        try:
            helper_bytes = io.BytesIO(await file2.read())
            email_mapping = build_email_mapping(helper_bytes)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"שגיאה בקריאת קובץ מיילים: {e}")

    # запускаем логики 1–7
    try:
        wb = process_workbook(wb, email_mapping=email_mapping)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"שגיאה בעיבוד הקובץ: {e}")

    # отдаём готовый Excel
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    headers = {
        "Content-Disposition": 'attachment; filename="גיול_אוטומציה_1-7.xlsx"'
    }

    return StreamingResponse(
        output,
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers=headers,
    )

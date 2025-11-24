import io
from collections import defaultdict
from datetime import datetime

import openpyxl
from openpyxl.styles import PatternFill, Alignment
import streamlit as st
import requests


# ========= הגדרות N8N =========
# כתובת Webhook אמיתית מ-node ה-Webhook ב-N8N (Production URL)
N8N_WEBHOOK_URL = "https://riseelena.app.n8n.cloud/webhook/e134717f-c0ad-4e29-a354-1b6edbe1d1ce"


# ---------- כלי עזר ----------

def parse_amount(val):
    """המרת ערך לסכום מספרי (float) עם טיפול בריק ומפרידי אלפים."""
    if val is None or val == "":
        raise ValueError("empty")
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    if s == "":
        raise ValueError("empty")
    s = s.replace(",", "")  # להסיר מפרידי אלפים
    return float(s)


def detect_headers(ws):
    """
    זיהוי שורת כותרות: מנסה שורה 1 ואז 2.
    מחזיר: (index של שורת כותרות, מילון {שם עמודה -> אינדקס עמודה})
    """
    candidates = [1, 2]
    chosen_row = None
    headers = {}

    for row_idx in candidates:
        row_cells = ws[row_idx]
        row_values = [str(c.value).strip() if c.value is not None else "" for c in row_cells]
        if any(v for v in row_values):
            tmp_headers = {str(c.value).strip(): c.column for c in ws[row_idx] if c.value}
            # עדיפות לשורה שיש בה גם "חשבון" וגם "חוב לחשבונית"
            if "חשבון" in tmp_headers and "חוב לחשבונית" in tmp_headers:
                chosen_row = row_idx
                headers = tmp_headers
                break
            # אם עוד לא בחרנו – נשמור ככותרת זמנית
            if not headers:
                chosen_row = row_idx
                headers = tmp_headers

    if chosen_row is None:
        chosen_row = 1
        headers = {str(c.value).strip(): c.column for c in ws[1] if c.value}

    return chosen_row, headers


# ---------- צבעים ----------

GREEN_RGB = "FF00FF00"   # ירוק
ORANGE_RGB = "FFFFA500"  # כת_
